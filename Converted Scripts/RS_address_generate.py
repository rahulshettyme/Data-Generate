# Author: Rahul Shetty
# Name: RS_address_generate
# Description: Fetch addresses/lat-lng from Excel, query Google Geocode, return address JSONs.

import requests, json, time

GOOGLE_API_KEY = "AIzaSyAwy--7hbQ9x-_rFT2lCi52o0rF0JvbA7E"
GEOCODE_URL = "https://maps.googleapis.com/maps/api/geocode/json"

# ==============================
# 🌍 Function: Get location details
# ==============================
def get_location_details(address_text=None, lat=None, lng=None):
    if address_text:
        params = {"address": address_text, "key": GOOGLE_API_KEY}
    elif lat is not None and lng is not None:
        params = {"latlng": f"{lat},{lng}", "key": GOOGLE_API_KEY}
    else:
        return None

    try:
        print(f"Calling Google Geocode API: {GEOCODE_URL} with params: { {k:(v if k!='key' else '***') for k,v in params.items()} }")
        response = requests.get(GEOCODE_URL, params=params)
        data = response.json()
        print(f"Google API Status: {response.status_code}")
        
        if response.status_code != 200 or not data.get("results"):
            print(f"Google API No Results or Error: {json.dumps(data)}")
            return None

        result = data["results"][0]
        
        # Safe component extraction
        address_components = result.get("address_components", [])
        def get_component(types):
            for comp in address_components:
                if any(t in comp["types"] for t in types):
                    return comp.get("long_name", "")
            return ""

        geometry = result.get("geometry", {}).get("location", {})
        res_lat = geometry.get("lat", lat)
        res_lng = geometry.get("lng", lng)

        address = {
            "country": get_component(["country"]),
            "formattedAddress": result.get("formatted_address", ""),
            "administrativeAreaLevel1": get_component(["administrative_area_level_1"]),
            "administrativeAreaLevel2": get_component(["administrative_area_level_2"]),
            "locality": get_component(["locality"]),
            "sublocalityLevel1": get_component(["sublocality_level_1"]),
            "sublocalityLevel2": get_component(["sublocality_level_2"]),
            "landmark": "",
            "postalCode": get_component(["postal_code"]),
            "houseNo": "",
            "buildingName": "",
            "placeId": result.get("place_id", ""),
            "latitude": res_lat,
            "longitude": res_lng
        }
        return address
    except Exception as e:
        print(f"Error fetching address: {e}")
        return None

def run(rows, token, env_config):
    processed_rows = []

    for row in rows:
        new_row = row.copy()
        
        # Helper to find value case-insensitively
        def get_val(keys):
            for k in keys:
                for row_k in row.keys():
                    if str(row_k).lower() == k.lower():
                        return row[row_k]
            return None

        address_text = get_val(["address", "location", "farmer address"])
        lat = get_val(["latitude", "lat"])
        lng = get_val(["longitude", "lng", "lon", "long"])

        try:
            time.sleep(0.1)
            
            address_info = None
            if lat and lng:
                print(f"Geocoding via Lat/Lng: {lat}, {lng}")
                address_info = get_location_details(lat=lat, lng=lng)
            elif address_text:
                print(f"Geocoding via Address: {address_text}")
                address_info = get_location_details(address_text=address_text)
            
            if address_info:
                new_row['Status'] = 'Pass'
                new_row['Formatted Address'] = address_info.get('formattedAddress')
                new_row['Latitude'] = address_info.get('latitude')
                new_row['Longitude'] = address_info.get('longitude')
                new_row['address_json'] = json.dumps(address_info)
                new_row['API_Response'] = new_row['address_json']
            else:
                # Fallback
                new_row['Status'] = 'Fail'
                new_row['API_Response'] = "Address Not Found"
                if address_text:
                    new_row['Formatted Address'] = address_text
                    new_row['Latitude'] = float(lat) if lat else 0
                    new_row['Longitude'] = float(lng) if lng else 0
                    fallback_json = {
                        "formattedAddress": address_text,
                        "latitude": new_row['Latitude'],
                        "longitude": new_row['Longitude']
                    }
                    new_row['address_json'] = json.dumps(fallback_json)

        except Exception as e:
            new_row['Status'] = 'Fail'
            new_row['API_Response'] = f"Error: {str(e)}"

        processed_rows.append(new_row)

    return processed_rows

# Optional standalone run (for local testing)
if __name__ == "__main__":
    import openpyxl
    def process_excel(file_path, sheet_name):
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]
        headers = {str(cell.value).strip(): idx+1 for idx, cell in enumerate(sheet[1]) if cell.value}
        data = []
        for r in range(2, sheet.max_row+1):
            row_data = {}
            for h, col_idx in headers.items():
                row_data[h] = sheet.cell(row=r, column=col_idx).value
            data.append(row_data)
        
        # Run logic
        results = run(data, "dummy_token", {})
        print(json.dumps(results, indent=2, ensure_ascii=False))

    # file_path = input("Enter Excel file path: ").strip()
    # sheet_name = input("Enter sheet name: ").strip()
    # process_excel(file_path, sheet_name)
    pass
